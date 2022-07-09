import openpyxl

excel_file = 'data/s3_임직원정보.xlsx'

# 엑셀파일 불러오기
wb = openpyxl.load_workbook(excel_file)

# 현재 활성화된 워크시트 선택
ws = wb.active

# 데이터 추가 1
ws['A1'] = '사번'
ws['B1'] = '이름'
ws['C1'] = '부서명'
ws['D1'] = '연락처'
ws['E1'] = '이메일'

# 데이터 추가 2
# ws.cell(row=2, column=1, value=101)
# ws.cell(row=2, column=2, value='김은민')
# ws.cell(row=2, column=3, value='영업기획부')
# ws.cell(row=2, column=4, value='010-1234-1111')
# ws.cell(row=2, column=5, value='aa@kjbank.com')

# 데이터 추가3 : list형태로 한번에 추가
ws.append([102, '박동현', 'IT기획부',     '010-1234-2222', 'bb@kjbank.com'])
ws.append([103, '이건호', '수도권전략부', '010-1234-3333', 'cc@kjbank.com'])
ws.append([104, '정은지', '리스크관리부', '010-1234-4444', 'dd@kjbank.com'])
ws.append([105, '최대훈', '각화동지점',   '010-1234-5555', 'ee@kjbank.com'])

# 시트명 변경
ws.title = '광주은행'

# 엑셀파일 저장
wb.save(excel_file)

print('프로그램 종료!')

