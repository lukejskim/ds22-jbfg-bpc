import openpyxl

# excel_file = './data/s2_시트생성삭제.xlsx'
excel_file = 'data/s3_임직원정보.xlsx'

# 엑셀파일 생성
wb = openpyxl.Workbook()

# 워크시트 생성
ws = wb.create_sheet('사용자시트')

# 모든 워크시트 이름 출력
print(wb.sheetnames)

# 워크시트 삭제
del wb['Sheet']

# 엑셀파일 저장
wb.save(excel_file)

print('프로그램 종료!')