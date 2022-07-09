import openpyxl

excel_file = 'data/JBFG_임직원.xlsx'

wb = openpyxl.load_workbook(excel_file)

# ws = wb.create_sheet('전북은행')
ws = wb['전북은행']

# 데이터 추가 1
ws['A1'] = '사번'
ws['B1'] = '이름'
ws['C1'] = '부서명'
ws['D1'] = '연락처'
ws['E1'] = '이메일'

# 데이터 추가 3
ws.append([ 201, '한혜형', '디지털영업팀', '010-5678-1111', 'aa@jbbank.com' ])
ws.append([ 202, '김영목', '카드사업부',   '010-5678-2222', 'bb@jbbank.com' ])
ws.append([ 203, '박성실', '데이터분석부', '010-5678-3333', 'cc@jbbank.com' ])
ws.append([ 204, '박요온', '데이터분석부', '010-5678-4444', 'dd@jbbank.com' ])
ws.append([ 205, '오승현', '지주',        '010-5678-5555', 'ee@jbbank.com' ])
ws.append([ 301, '김진수', '전략기획본부', '010-5678-6666', 'zz@jbbank.com' ])

wb.save(excel_file)

print("프로그램 종료!!")