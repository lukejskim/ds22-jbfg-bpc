import openpyxl
from openpyxl.utils import get_column_letter

excel_file = './data/JBFG_임직원.xlsx'

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# 데이터조회
# A1 = ws['A1'].value
# print(" A1 = {} ".format(A1))
#
# B1 = ws.cell(column=2, row=1).value
# print(" B1 = {} ".format(B1))
#
# B2 = ws.cell(column=2, row=2).value
# print(" B2 = {} ".format(B2))


rows = ws.rows
print('rows : {}, {}'.format(type(rows), rows))

arr_row = list()
cnt = 0
for row in rows:
    arr_row.append(row)
    for cell in row:
        val = cell.value
        print(val, end='\t')
    print()

# print(arr_row)

# print('-'*50)
# print(arr_row[0][0].value)
# print(arr_row[0][1].value)
# print(arr_row[0][2].value)


# for r_no in range(1, 6):
#     # print(r_no)
#     for c_no in range(1,7):
#         cell_loc = "{}{}".format(get_column_letter(c_no), r_no)
#         print(ws[cell_loc].value, end='  ')
#         # cell_loc = "{}{} ".format(get_column_letter(c_no), r_no)
#         # print(ws[cell_loc], end='  ')
#     print()


wb.close()

print("프로그램 종료!!")



"""
최종출력

사번	이름	부서명	연락처	이메일
101	김은민	영업기획부	010-1234-1111	aa@kjbank.com
102	박동현	IT기획부	010-1234-2222	bb@kjbank.com
103	이건호	수도권전략부	010-1234-3333	cc@kjbank.com
104	정은지	리스크관리부	010-1234-4444	dd@kjbank.com
105	최대훈	각화동지점	010-1234-5555	ee@kjbank.com


"""

