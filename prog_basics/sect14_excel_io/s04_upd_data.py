import openpyxl

excel_file = 'data/s3_임직원정보.xlsx'

# 엑셀파일 불러오기
wb = openpyxl.load_workbook(excel_file)

# 워크시트 선택
ws = wb['전북은행']

# 데이터 수정
ws['C7'] = '빅파이크래프트'
ws['D7'] = '010-5670-3847'
ws['E7'] = 'bigpycraft@gmail.com'

# 워크시트 선택
ws = wb['광주은행']

# 데이터 추가3 : list형태로 한번에 추가
ws.append([102, '박동현', 'IT기획부',     '010-1234-2222', 'bb@kjbank.com'])
ws.append([103, '이건호', '수도권전략부', '010-1234-3333', 'cc@kjbank.com'])
ws.append([104, '정은지', '리스크관리부', '010-1234-4444', 'dd@kjbank.com'])
ws.append([105, '최대훈', '각화동지점',   '010-1234-5555', 'ee@kjbank.com'])

# 엑셀파일 저장
wb.save(excel_file)

print('프로그램 종료!')

