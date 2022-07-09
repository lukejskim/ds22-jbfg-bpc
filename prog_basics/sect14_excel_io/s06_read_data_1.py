import openpyxl

excel_file = 'data/s3_임직원정보.xlsx'

# 엑셀파일 불러오기
wb = openpyxl.load_workbook(excel_file)


# 현재 활성화된 워크시트 선택
ws = wb.active

# data_only=True 옵션 : 셀값이 함수/수식일 경우 결과값만 가져옴
# wb = openpyxl.load_workbook(excel_file, data_only=True)
# ws = wb['광주은행']
# ws = wb['전북은행']

# 데이터 조회
A1 = ws['A1'].value
print("ws['A1'] = {}".format(A1))

B1 = ws.cell(column=2, row=1).value
print("ws['B1'] = {}".format(B1))

B2 = ws.cell(column=2, row=2).value
print("ws['B2'] = {}".format(B2))

# 엑셀파일 닫기
wb.close()

print('프로그램 종료!')

