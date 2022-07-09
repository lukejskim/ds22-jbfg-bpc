import openpyxl
from openpyxl.utils import get_column_letter

excel_file = 'data/s3_임직원정보.xlsx'

# 엑셀파일 불러오기
wb = openpyxl.load_workbook(excel_file)

# 현재 활성화된 워크시트 선택
ws = wb.active

# data_only=True 옵션 : 셀값이 함수/수식일 경우 결과값만 가져옴
# wb = openpyxl.load_workbook(excel_file, data_only=True)
# ws = wb['광주은행']
# ws = wb['전북은행']

# 모든 행(가로) 접근
rows = ws.rows
print('rows : {}, {}'.format(type(rows), rows))
# print('rows :', type(rows), rows)

# 행 데이터 확인
arr_row = list()
cnt = 0
for row in rows:
    arr_row.append(row)
    for cell in row:
        val = cell.value
        print(val, end='\t | ')
    print()

print(arr_row[0][0].value)
print(arr_row[0][1].value)
print(arr_row[0][2].value)


print('-'*50)

# 엑셀파일 닫기
wb.close()

print('프로그램 종료!')

