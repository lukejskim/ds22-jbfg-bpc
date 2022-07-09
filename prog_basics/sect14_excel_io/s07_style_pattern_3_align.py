import openpyxl
from openpyxl.styles import PatternFill, colors, Font, Alignment

excel_file = 'data/s3_임직원정보.xlsx'

# 엑셀파일 불러오기
wb = openpyxl.load_workbook(excel_file)

# 워크시트 선택
ws = wb.active
# ws = wb['광주은행']
# ws = wb['전북은행']

# 배경 색상
pattern_red    = PatternFill(start_color="FF0000", fill_type="solid")
pattern_yellow = PatternFill(start_color="FFFF00", fill_type="solid")
pattern_blue   = PatternFill(start_color=colors.BLUE, fill_type="solid")

ws.cell(1, 2).fill = pattern_red
ws.cell(5, 3).fill = pattern_yellow
ws.cell(4, 5).fill = pattern_blue

# 폰트
font_20 = Font(name="나눔고딕", size=20, color=colors.BLUE)
ws.cell(1, 1).font = font_20
ws.cell(2, 2).font = font_20

# 정렬
align_center = Alignment(horizontal="center", vertical="center")
for idx in range(1, 6):
    ws.cell(1, idx).alignment = align_center

for idx in range(1, 7):
    ws.cell(idx, 2).alignment = align_center


# 엑셀파일 저장
wb.save(excel_file)

print('프로그램 종료!')

