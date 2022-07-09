import openpyxl
from openpyxl.utils import get_column_letter

excel_file = 'data/s3_임직원정보.xlsx'

# 엑셀파일 불러오기
wb = openpyxl.load_workbook(excel_file)


# 현재 활성화된 워크시트 선택
ws = wb.active

# data_only=True 옵션 : 셀값이 함수/수식일 경우 결과값만 가져옴
# wb = openpyxl.load_workbook(excel_file, data_only=True)
# ws = wb['광주은행']
# ws = wb['전북은행']

# 데이터 조회 : N개
A1B2 = ws['A1:B2']
print("ws['A1:B2'] = {}".format(A1B2))

A1E1 = ws['A1:E1']
print("ws['A1:E1'] = {}".format(A1E1))
print(A1E1[0][0].value)
print(A1E1[0][1].value)
print(A1E1[0][2].value)

A1E6 = ws['A1:E6']
s_data = A1E6
print("1. {} : {}".format(type(s_data), s_data))
print("2. {} : {}".format(type(s_data[0]), s_data[0]))
print("3. {} : {}".format(type(s_data[0][0]), s_data[0][0]))
print("4. {} : {}".format(type(s_data[0][0].value), s_data[0][0].value))

cell =  s_data[0][0]
print("5. {} : {}".format('s_data', len(s_data)))
print("6. {} : {}".format('s_data[0]', len(s_data[0])))
# print("7. {} : {}".format('s_data[0][0]', len(s_data[0][0])))


for row in range(len(s_data)):
    # print("{} : {}".format(type(row), row))
    for col in range(len(s_data[row])):
        cell_val = s_data[row][col].value
        x_col, x_row = get_column_letter(col+1), row+1
        cell_loc = "{}{}".format(x_col, x_row)
        print("{} : {}".format(cell_loc, cell_val), end='\n')
        # print("{} : {}".format(row, col))
        # print(cell, end='\t')
    print()


# for row in A1E1:
#     print("{} : {}".format(type(row), row))
#     for column in row:
#         cell_loc = ''
#         cell_val = column.value
#         print("{} : {}".format(cell_loc, cell_val))
        # print("{} : {}".format(type(column), column.value))
        # cell = A1E1[rows][cols].value
        # print(cell, end='\t')
print()



#
# A1 = ws['A1'].value
# print("ws['A1'] = {}".format(A1))
#
# B1 = ws.cell(column=2, row=1).value
# print("ws['B1'] = {}".format(B1))
#
# B2 = ws.cell(column=2, row=2).value
# print("ws['B2'] = {}".format(B2))

# 엑셀파일 닫기
wb.close()

print('프로그램 종료!')

